#!/usr/bin/env python3
"""
Genera proyectos_data.json a partir de:
  - PEN_y_CD.xlsx      (planilla madre de proyectos de ley del Senado, TIPO == 'PL')
  - fuentes/*.md        (notas por proyecto: resumen del mensaje, del expediente,
                          reuniones de comisión, orden del día y sanción)

Uso:
    python3 data_builder.py
    python3 data_builder.py --xlsx otra_planilla.xlsx --fuentes otra_carpeta --out otro.json

Al agregar un .md nuevo a fuentes/, o al reemplazar el Excel, correr de nuevo
este script regenera proyectos_data.json completo.
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

from md_parser import parse_md_file, strip_markup, sentence_case

HERE = Path(__file__).resolve().parent

# Títulos "lindos" + nombre corto para tarjetas, curados a mano
# (clave: (ORIGEN, NRO, AÑO) -> (titulo completo, nombre corto))
TITULOS_CURADOS = {
    ("CD", 1, "2026"): ("Ley de Hojarasca: reforma para la desregulación y simplificación normativa", "Hojarasca"),
    ("PE", 13, "2026"): ("Ley de Inviolabilidad de la Propiedad Privada y desregulación territorial", "Inviolabilidad de la propiedad privada"),
    ("PE", 186, "2026"): ("Derogación de la Ley de Etiquetado Frontal", "Etiquetado frontal"),
    ("PE", 7, "2026"): ("Reforma de la Carta Orgánica del Banco Central", "Reforma carta orgánica BCRA"),
    ("CD", 3, "2026"): ("Reforma energética: readecuación de subsidios de zona fría y saneamiento del Mercado Eléctrico Mayorista", "Zonas frías"),
    ("PE", 193, "2026"): ("Reforma de la Ley General de Sociedades y economía digital", "Reforma ley general de sociedades"),
    ("CD", 6, "2026"): ("Régimen de Incentivo para Grandes Inversiones en Nuevas Industrias (Súper RIGI)", "Súper RIGI"),
    ("PE", 99, "2026"): ("Reforma de la Ley Nacional de Salud Mental", "Salud mental"),
    ("PE", 102, "2026"): ("Reforma electoral integral", "Reforma electoral"),
    ("PE", 100, "2026"): (None, "Pensiones por invalidez"),
    ("PE", 185, "2026"): (None, "Ludopatía y juegos de azar"),
}

# Claves de expediente que el parseo automático de la línea EXPEDIENTE del .md
# no puede resolver bien (identificadores duales/atípicos) -> se fuerza a mano.
CLAVE_OVERRIDE_POR_ARCHIVO = {
    "Reforma_Energetica_ZonasFrias.md": ("CD", 3, "2026"),
    "Super_RIGI.md": ("CD", 6, "2026"),
}

# Proyectos de los que tenemos nota .md pero NO están en la planilla del Senado
# porque todavía tramitan en Diputados. Se arman a mano con lo que hay en la nota;
# faltan datos (url, fechas de giro a comisión) hasta que tengamos su fuente propia.
PROYECTOS_SIN_EXCEL = {
    "Reforma_Carta_Organica_BCRA.md": {
        "expediente": "7-PE/26 (Diputados)",
        "url": None,
        "camara_actual": "Diputados",
        "categoria": "en_comision",
        "titulo": "Reforma de la Carta Orgánica del Banco Central",
        "nombre_corto": "Reforma carta orgánica BCRA",
        "faltantes": ["url del expediente", "fechas de giro a comisión", "N° de expediente definitivo en Diputados"],
    },
    "Lobby.md": {
        "expediente": "Mensaje PEN N° 178/26 (Diputados)",
        "url": None,
        "camara_actual": "Diputados",
        "categoria": "en_comision",
        "titulo": "Régimen de Gestión de Intereses (Ley de Lobby)",
        "nombre_corto": "Lobby",
        "faltantes": ["N° de expediente en Diputados", "url del expediente", "fechas de giro a comisión"],
    },
    "Inocencia_Fiscal.md": {
        "expediente": "Mensaje PEN N° 223/26 (Diputados)",
        "url": None,
        "camara_actual": "Diputados",
        "categoria": "en_comision",
        "titulo": "Simplificación tributaria y actualización del régimen de sanciones laborales (Inocencia Fiscal 2)",
        "nombre_corto": "Inocencia fiscal",
        "faltantes": ["N° de expediente en Diputados", "url del expediente", "fechas de giro a comisión"],
    },
}


def detectar_camara(texto, default):
    """Busca en el texto una mención explícita a la cámara (Senado / Diputados / HCDN / HSN)
    y si no encuentra ninguna, devuelve la cámara por defecto que se le pase."""
    if not texto:
        return default
    lower = texto.lower()
    if "diputados" in lower or "hcdn" in lower:
        return "Diputados"
    if "senado" in lower or "hsn" in lower:
        return "Senado"
    return default


def etiquetar_camaras(nota_dict, origen, badge_camara, camara_reunion_default):
    """Agrega el campo 'camara' a reunion_comision, orden_del_dia_detalle y
    sancion_detalle indicando explícitamente a qué cámara refiere esa información."""
    # si vino de Diputados (badge HCDN) el OD/la sanción documentados en la nota son
    # los que ya se emitieron en Diputados; si el Senado ya lo aprobó (badge HSN) son
    # los del propio Senado.
    if badge_camara == "1/2 sanción HCDN":
        default_od_sancion = "Diputados"
    elif badge_camara == "1/2 sanción HSN":
        default_od_sancion = "Senado"
    else:
        default_od_sancion = camara_reunion_default

    if nota_dict.get("reunion_comision"):
        rc = nota_dict["reunion_comision"]
        rc["camara"] = detectar_camara(rc.get("comision", ""), camara_reunion_default)
    if nota_dict.get("orden_del_dia_detalle"):
        od = nota_dict["orden_del_dia_detalle"]
        texto_od = " ".join(od.get("firmantes", []))
        od["camara"] = detectar_camara(texto_od, default_od_sancion)
    if nota_dict.get("sancion_detalle"):
        sd = nota_dict["sancion_detalle"]
        texto_sancion = " ".join(sd.get("posturas_principales", []))
        sd["camara"] = detectar_camara(texto_sancion, default_od_sancion)
    return nota_dict


def clean_titulo_desde_caratula(caratula):
    """Fallback: limpia el prefijo 'MENSAJE N° X/Y Y ' / 'PROYECTO DE LEY (EN REVISION) '
    y arma un título legible en formato oración (best-effort, sin tildes que la
    planilla ya no tenía)."""
    t = caratula or ""
    t = re.sub(r'^MENSAJE\s*N[°ºo]?\s*\d+\s*/\s*\d+\s*Y\s+', '', t, flags=re.IGNORECASE)
    t = re.sub(r'^PROYECTO DE LEY(\s+EN REVISION)?\s+', '', t, flags=re.IGNORECASE)
    t = t.rstrip(". ").strip().strip('"').strip()

    if re.match(r'^QUE\b', t, re.IGNORECASE):
        t = "PROYECTO DE LEY " + t
    elif re.match(r'^(SOBRE|GENERAL DE|DE|CONTRA|PARA)\b', t, re.IGNORECASE):
        t = "LEY " + t

    return sentence_case(t)


def parse_od_field(raw):
    if not raw:
        return None
    text = str(raw).strip()
    if not re.match(r'^\d+\s+\d{4}\b', text):
        return None
    m = re.match(r'^(\d+)\s+(\d{4})', text)
    return {"numero": m.group(1), "anio": m.group(2)}


def parse_sancion_field(raw):
    if not raw:
        return None
    text = str(raw).strip()
    if not text.upper().startswith("AP"):
        return None
    m = re.search(r'AP\s+(\d{2}/\d{2}/\d{4})', text.upper())
    return m.group(1) if m else None


def extraer_fecha(raw):
    if not raw:
        return None
    m = re.search(r'\d{2}/\d{2}/\d{4}', str(raw))
    return m.group(0) if m else None


def extraer_fechas(raw):
    if not raw:
        return []
    return re.findall(r'\d{2}/\d{2}/\d{4}', str(raw))


def tiene_fecha_real(raw):
    if not raw:
        return False
    return bool(re.search(r'\d{2}/\d{2}/\d{4}', str(raw)))


def build(xlsx_path, fuentes_dir):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    # --- notas .md, indexadas por (origen, nro, anio) ---
    notas_por_clave = {}
    notas_sin_match = []
    notas_por_archivo = {}
    for md_path in sorted(fuentes_dir.glob("*.md")):
        parsed = parse_md_file(md_path)
        notas_por_archivo[md_path.name] = parsed

        override = CLAVE_OVERRIDE_POR_ARCHIVO.get(md_path.name)
        claves = [override] if override else parsed["expediente_claves"]
        if not claves:
            if md_path.name not in PROYECTOS_SIN_EXCEL:
                notas_sin_match.append((md_path.name, "no se pudo interpretar la línea EXPEDIENTE"))
            continue
        for clave in claves:
            notas_por_clave[clave] = parsed

    proyectos = []
    excluidos_ya_ley = []
    od_sin_parsear = []
    claves_usadas = set()

    for row in ws.iter_rows(min_row=2):
        if row[col["TIPO"]].value != "PL":
            continue

        origen = row[col["ORIGEN"]].value
        nro = int(row[col["NRO."]].value)
        anio_f = row[col["AÑO"]].value
        anio = str(int(anio_f)) if anio_f else "2026"
        yy = anio[-2:]
        clave = (origen, nro, anio)

        ley_fecha_raw = row[col["LEY – FECHA"]].value
        if tiene_fecha_real(ley_fecha_raw):
            excluidos_ya_ley.append({
                "expediente": f"{origen}-{nro}/{yy}",
                "caratula": (row[col["CARÁTULA"]].value or "").strip(),
                "ley_fecha": extraer_fecha(ley_fecha_raw),
            })
            continue

        caratula = row[col["CARÁTULA"]].value or ""
        nota = notas_por_clave.get(clave)
        if nota:
            claves_usadas.add(nota["archivo"])
        curado = TITULOS_CURADOS.get(clave)
        titulo = (curado[0] if curado and curado[0] else None) or clean_titulo_desde_caratula(caratula)
        nombre_corto = (curado[1] if curado else None) or titulo

        od_raw = row[col["ORDEN DEL DÍA"]].value
        od = parse_od_field(od_raw)
        if od_raw and not od and str(od_raw).strip() not in ("", "-"):
            od_sin_parsear.append({"expediente": f"{origen}-{nro}/{yy}", "valor": od_raw})

        sancion_fecha = parse_sancion_field(row[col["SANCIONES/SITUACIÓN EXP"]].value)

        # --- categoría ---
        if sancion_fecha:
            categoria = "media_sancion"
        elif od:
            categoria = "con_od"
        else:
            categoria = "en_comision"

        # --- badge de cámara y cámara donde está el expediente hoy ---
        if origen == "CD":
            badge_camara = "1/2 sanción HCDN"
            camara_actual = "Senado"  # vino de Diputados con 1/2 sanción, ahora se trata en el Senado
        elif sancion_fecha:
            badge_camara = "1/2 sanción HSN"
            camara_actual = "Diputados"  # el Senado ya lo aprobó, pasó a la otra cámara
        else:
            badge_camara = None
            camara_actual = "Senado"  # origen PE, todavía sin 1/2 sanción: sigue en el Senado

        comisiones = []
        for i in (1, 2, 3):
            nombre = row[col[f"COMISION{i}"]].value
            if not nombre:
                continue
            comisiones.append({
                "comision": nombre,
                "fecha_ingreso": row[col[f"FECHA_INGRESO{i}"]].value.strftime("%d/%m/%Y") if row[col[f"FECHA_INGRESO{i}"]].value else None,
                "fecha_egreso": row[col[f"FECHA_EGRESO{i}"]].value.strftime("%d/%m/%Y") if row[col[f"FECHA_EGRESO{i}"]].value else None,
            })

        url = f"https://www.senado.gob.ar/parlamentario/comisiones/verExp/{nro}.{yy}/{origen}/PL"

        proyecto = {
            "expediente": f"{origen}-{nro}/{yy}",
            "origen": origen,
            "autor": "Poder Ejecutivo Nacional",
            "url": url,
            "titulo": titulo,
            "nombre_corto": nombre_corto,
            "caratula_original": caratula.strip(),
            "categoria": categoria,
            "badge_camara": badge_camara,
            "camara_actual": camara_actual,
            "tramite": {
                "ingreso": extraer_fecha(row[col["MESA DE ENTRADAS"]].value),
                "comisiones": comisiones,
                "fecha_ingreso_dictamen": extraer_fechas(row[col["FECHA INGRESO DICTAMEN"]].value),
                "orden_del_dia": od,
                "sancion_fecha": sancion_fecha,
            },
            "nota": None,
        }

        if nota:
            proyecto["nota"] = etiquetar_camaras({
                "resumen_mensaje_pen": nota["resumen_mensaje_pen"],
                "resumen_expediente": nota["resumen_expediente"],
                "reunion_comision": nota["reunion_comision"],
                "orden_del_dia_detalle": nota["orden_del_dia_md"],
                "sancion_detalle": nota["sancion"],
            }, origen, badge_camara, "Senado")

        proyectos.append(proyecto)

    # --- proyectos armados a mano porque no están en la planilla del Senado ---
    proyectos_manuales_info = []
    for archivo, manual in PROYECTOS_SIN_EXCEL.items():
        nota = notas_por_archivo.get(archivo)
        if not nota:
            continue
        claves_usadas.add(archivo)
        proyectos.append({
            "expediente": manual["expediente"],
            "origen": None,
            "autor": "Poder Ejecutivo Nacional",
            "url": manual["url"],
            "titulo": manual["titulo"],
            "nombre_corto": manual["nombre_corto"],
            "caratula_original": None,
            "categoria": manual["categoria"],
            "badge_camara": None,
            "camara_actual": manual["camara_actual"],
            "tramite": {
                "ingreso": None,
                "comisiones": [],
                "fecha_ingreso_dictamen": [],
                "orden_del_dia": None,
                "sancion_fecha": None,
            },
            "nota": etiquetar_camaras({
                "resumen_mensaje_pen": nota["resumen_mensaje_pen"],
                "resumen_expediente": nota["resumen_expediente"],
                "reunion_comision": nota["reunion_comision"],
                "orden_del_dia_detalle": nota["orden_del_dia_md"],
                "sancion_detalle": nota["sancion"],
            }, None, None, manual["camara_actual"]),
            "sin_datos_de_planilla": True,
        })
        proyectos_manuales_info.append({
            "expediente": manual["expediente"],
            "archivo": archivo,
            "faltantes": manual["faltantes"],
        })

    vistos = set()
    notas_sin_excel_dedup = []
    for nota in notas_por_clave.values():
        if nota["archivo"] in claves_usadas or nota["archivo"] in vistos:
            continue
        vistos.add(nota["archivo"])
        notas_sin_excel_dedup.append({"archivo": nota["archivo"], "expediente": nota["expediente_raw"]})

    resultado = {
        "generado_al": None,
        "total": len(proyectos),
        "proyectos": proyectos,
    }
    log = {
        "excluidos_ya_ley": excluidos_ya_ley,
        "orden_del_dia_sin_parsear": od_sin_parsear,
        "notas_sin_match_en_excel": notas_sin_excel_dedup,
        "notas_con_expediente_ilegible": notas_sin_match,
        "proyectos_manuales_incompletos": proyectos_manuales_info,
    }
    return resultado, log


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", default=HERE / "PEN_y_CD.xlsx")
    parser.add_argument("--fuentes", default=HERE / "fuentes")
    parser.add_argument("--out", default=HERE / "proyectos_data.json")
    args = parser.parse_args()

    xlsx_path, fuentes_dir, out_path = Path(args.xlsx), Path(args.fuentes), Path(args.out)
    if not xlsx_path.exists():
        sys.exit(f"ERROR: no se encuentra el archivo fuente: {xlsx_path}")
    if not fuentes_dir.exists():
        sys.exit(f"ERROR: no se encuentra la carpeta de notas: {fuentes_dir}")

    resultado, log = build(xlsx_path, fuentes_dir)

    import datetime
    resultado["generado_al"] = datetime.date.today().isoformat()

    out_path.write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {resultado['total']} proyectos pendientes escritos en {out_path}")
    print()

    if log["excluidos_ya_ley"]:
        print("--- Excluidos: ya sancionados como ley ---")
        for item in log["excluidos_ya_ley"]:
            print(f"  {item['expediente']} (ley del {item['ley_fecha']}): {item['caratula'][:90]}")
        print()

    if log["orden_del_dia_sin_parsear"]:
        print("--- AVISO: Orden del Día con formato no interpretable (revisar a mano) ---")
        for item in log["orden_del_dia_sin_parsear"]:
            print(f"  {item['expediente']}: {item['valor']!r}")
        print()

    if log["notas_sin_match_en_excel"]:
        print("--- AVISO: notas .md sin fila en el Excel y sin entrada manual (agregar a PROYECTOS_SIN_EXCEL o revisar clave) ---")
        for item in log["notas_sin_match_en_excel"]:
            print(f"  {item['archivo']} (expediente detectado: {item['expediente']})")
        print()

    if log["notas_con_expediente_ilegible"]:
        print("--- AVISO: no se pudo leer la línea EXPEDIENTE de estas notas ---")
        for archivo, motivo in log["notas_con_expediente_ilegible"]:
            print(f"  {archivo}: {motivo}")
        print()

    if log["proyectos_manuales_incompletos"]:
        print("--- AVISO: proyectos cargados a mano (no están en la planilla del Senado) — falta completar ---")
        for item in log["proyectos_manuales_incompletos"]:
            print(f"  {item['expediente']} ({item['archivo']}): falta " + ", ".join(item["faltantes"]))
        print()

    proyectos_con_nota = sum(1 for p in resultado["proyectos"] if p["nota"])
    print(f"Proyectos con nota .md cargada: {proyectos_con_nota} / {resultado['total']}")


if __name__ == "__main__":
    main()
